import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return(sheet.max_row)

def getColumnCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return(sheet.max_column)

def readData(file,SheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return(sheet.cell(rowNum,colNum).value)

def writeData(file,SheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(rowNum,colNum).value=data
    workbook.save(file)

def fillGreenColour(file,SheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    greenFill=PatternFill(start_color='00BC55',end_color='00BC55',fill_type='solid')
    sheet.cell(rowNum,colNum).fill=greenFill
    workbook.save(file)

def fillRedColour(file,SheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    redFill=PatternFill(start_color='FD492B',end_color='FD492B',fill_type='solid')
    sheet.cell(rowNum,colNum).fill=redFill
    workbook.save(file)